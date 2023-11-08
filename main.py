import os
import tkinter as tk
from threading import Thread
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinterdnd2 import TkinterDnD, DND_FILES
from vlookup import perform_vlookup
import pygame


# plays last Christmas when the gui is opened lol
def play_background_music():
    pygame.mixer.init()
    pygame.mixer.music.load('images-videos/last_christmas.mp3')
    pygame.mixer.music.play(-1)  # Play the music, -1 means play indefinitely in loop
    pass


music_thread = Thread(target=play_background_music, daemon=True)
music_thread.start()


def sort_excel(file_path, sort_columns, ascending_order, file_type=""):
    if not sort_columns:
        messagebox.showerror("Error", "No columns selected for sorting.")
        return

    try:
        # Read the Excel file into a pandas DataFrame
        df = pd.read_excel(file_path)

        # If 'SND Cost' is one of the sort columns, convert it to numeric
        if 'SND Cost' in sort_columns:
            df['SND Cost'] = pd.to_numeric(df['SND Cost'], errors='coerce')

        # If 'VPC Cost' is one of the sort columns, convert it to numeric
        if 'VPC Cost' in sort_columns:
            df['VPC Cost'] = pd.to_numeric(df['VPC Cost'], errors='coerce')

        # Sort the DataFrame based on the selected columns
        df = df.sort_values(by=sort_columns, ascending=ascending_order)

        # Save the sorted DataFrame back to the Excel file
        df.to_excel(file_path, index=False)

        # messagebox.showinfo("Success!", f"Success! {file_type} file sorted and saved successfully.")
        # Don't use we don't need both success messages


    except Exception as e:
        messagebox.showerror("Error", str(e))


def process_file(file_path):
    filename = os.path.basename(file_path).lower()
    if 'award' in filename:
        sort_excel(file_path, ['Product ID', 'Award Cust ID'], [True, False], "Award")
    elif 'backlog' in filename:
        sort_excel(file_path, ['Product ID', 'Backlog Entry'], [True, False], "Backlog")
    elif 'sales' in filename:
        sort_excel(file_path, ['Product ID', 'Last Ship Date'], [True, False], "Sales History")
    elif 'snd' in filename:
        sort_excel(file_path, ['Product ID', 'SND Cost'], [True, True], "Ship & Debit")
    elif 'vpc' in filename:
        sort_excel(file_path, ['PART ID', 'VPC Cost'], [True, False], "VPC")


def merge_files_and_create_lost_items(folder_path):
    print("Merging files and creating lost items...")

    # Define the order and corresponding keywords in which the files should be merged
    merge_order = [
        ('Awards', 'award'),
        ('Backlog', 'backlog'),
        ('Sales History', 'sales'),
        ('SND', 'snd'),
        ('VPC', 'vpc')
    ]

    # Find all Excel files in the folder
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    # Extracting contract files based on naming convention and dates in filenames
    contract_files = sorted([f for f in excel_files if 'CT ACTIVE CONTRACTS' in f],
                            key=lambda x: x[-12:])  # Sort by date at the end of filename

    if len(contract_files) < 2:
        messagebox.showerror("Error", "Not enough contract files found. Need current and previous week's files.")
        return

    # Assume the last two files are current and previous based on naming convention
    current_contract_path = os.path.join(folder_path, contract_files[-1])
    prev_contract_path = os.path.join(folder_path, contract_files[-2])

    # Load the workbooks and dataframes for the current and previous contracts
    active_workbook = load_workbook(current_contract_path)
    active_df = pd.read_excel(current_contract_path, sheet_name=0, header=1)
    prev_df = pd.read_excel(prev_contract_path, sheet_name=0, header=0)

    # Now we have the dataframes, we need to check if 'IPN' column exists
    if 'IPN' not in prev_df.columns or 'IPN' not in active_df.columns:
        messagebox.showerror("Error", "'IPN' column is missing in one of the contract files.")
        return

    # Create a 'Lost Items' sheet, checked the IPN from both sheets
    lost_items_sheet = active_workbook.create_sheet('Lost Items')
    lost_items_df = prev_df[~prev_df['IPN'].isin(active_df['IPN'])]
    if not lost_items_df.empty:
        for r in dataframe_to_rows(lost_items_df, index=False, header=True):
            lost_items_sheet.append(r)

    # Create a 'Prev Contract' sheet
    prev_contract_sheet = active_workbook.create_sheet('Prev Contract')
    for r in dataframe_to_rows(prev_df, index=False, header=True):
        prev_contract_sheet.append(r)

    # Loop through the files in the specified order and merge them
    for sheet_name, keyword in merge_order:
        for file_name in excel_files:
            if keyword.lower() in file_name.lower() and file_name not in contract_files:
                file_path = os.path.join(folder_path, file_name)
                data_df = pd.read_excel(file_path, sheet_name=0)
                new_sheet = active_workbook.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(data_df, index=False, header=True):
                    new_sheet.append(r)
                break  # Once the correct file is found and processed, break the loop

    # Save the active workbook with the new sheets
    try:
        active_workbook.save(current_contract_path)
        messagebox.showinfo("Success!", "Files merged and sheets created successfully in order.")
    except Exception as e:
        messagebox.showerror("Error", str(e))


def on_drop(event):
    folder_path = event.data
    if os.path.isdir(folder_path):
        process_folder(folder_path)
        messagebox.showinfo("Complete", "Files have been processed.")
    else:
        messagebox.showerror("Error", "Please drop a folder, not files.")


# path to process folder, drop-off created

def process_folder(folder_path):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path) and filename.endswith('.xlsx'):
            process_file(file_path)
    # After all files have been processed, proceed to merge and create the lost items sheet
    merge_files_and_create_lost_items(folder_path)


def setup_gui(root):
    style = ttk.Style()
    root.configure(bg="white")
    style.configure("TButton", font=("Roboto", 16, "bold"), width=40, height=40)
    style.map("TButton", foreground=[('active', 'white')], background=[('active', '#007BFF')])

    title_label = ttk.Label(root, text="Welcome Partnership Team!",
                            font=("Segoe UI", 36, "underline"), background="white", foreground="#103d81")
    title_label.pack(pady=20)

    run_queries_button = ttk.Button(root, text="Run Queries", command=print("run queries"), style="TButton")
    run_queries_button.pack(pady=20)

    description_label = ttk.Label(root,
                                  text="This tool allows you to sort your Excel files for our Creation \n"
                                       "Contact Automatically once saved in your desired folder.\n "
                                       "Please drop your folder below ↓",
                                  font=("Roboto", 18), background="white", anchor="center",
                                  justify="center")
    description_label.pack(pady=20)

    # Create a label that will act as the drop area
    drop_area = tk.Label(root, text='Drop Folder Here', pady=50, padx=50, relief="groove", borderwidth=2,
                         font=("Roboto", 20))
    drop_area.pack(pady=30, padx=30)  # Add some padding to make the border visible

    perform_vlookup_button = ttk.Button(root, text="Perform VLook-Up to new file",
                                        command=lambda: perform_vlookup(perform_vlookup_button), style="TButton")
    perform_vlookup_button.pack(pady=20)

    # Register label as a drop target
    drop_area.drop_target_register(DND_FILES)
    drop_area.dnd_bind('<<Drop>>', on_drop)


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    root.title('File Processor')
    root.geometry('700x600')  # Set the window size to be bigger
    setup_gui(root)
    root.mainloop()
